#!/usr/bin/env python3
"""Shrink a national CMS file to Kentucky before it ever needs uploading.

The raw CMS downloads run from tens of megabytes to well over a gigabyte because they
carry every contract, plan and county in the country. Kentucky is roughly 2 to 5 percent
of that. Run this on the downloaded file and you get a small CSV.

Usage
    python3 docs/filter-cms-file.py <downloaded-file> [output.csv] [--state KY]

Handles
    .csv   any CMS csv (penetration, contract info, enrollment)
    .xlsb  the landscape files, needs:  pip install pyxlsb
    .xlsx  needs:  pip install openpyxl

It finds the state column itself, so it works across the different CMS layouts, and it
streams rather than loading the file into memory.
"""
import sys, os, csv, re

STATE_HINTS = ("state territory abbreviation", "state abbreviation", "state name",
               "state", "statename", "st")
FULL = {"KY": "kentucky"}


def state_col(header, want):
    """Pick the column most likely to hold the state, preferring abbreviation columns."""
    low = [str(h or "").strip().lower() for h in header]
    for hint in STATE_HINTS:
        for i, h in enumerate(low):
            if h == hint:
                return i
    for i, h in enumerate(low):
        if "state" in h:
            return i
    return None


def matches(value, want):
    v = str(value or "").strip().lower()
    return v == want.lower() or v == FULL.get(want.upper(), "\0")


def rows_from_csv(path):
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, newline="", encoding=enc) as f:
                for row in csv.reader(f):
                    yield row
            return
        except UnicodeDecodeError:
            continue
    raise SystemExit("Could not decode the file as UTF-8, cp1252 or latin-1.")


def rows_from_xlsb(path):
    try:
        from pyxlsb import open_workbook
    except ImportError:
        raise SystemExit("This is an .xlsb file. Install the reader first:\n\n    pip install pyxlsb\n")
    with open_workbook(path) as wb:
        with wb.get_sheet(1) as sh:
            for row in sh.rows():
                yield [c.v for c in row]


def rows_from_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("This is an .xlsx file. Install the reader first:\n\n    pip install openpyxl\n")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for row in wb[wb.sheetnames[0]].iter_rows(values_only=True):
        yield list(row)


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    want = "KY"
    for a in sys.argv[1:]:
        if a.startswith("--state"):
            want = a.split("=", 1)[1] if "=" in a else "KY"
    if not args:
        print(__doc__)
        raise SystemExit(2)

    src = args[0]
    if not os.path.exists(src):
        raise SystemExit(f"No such file: {src}")
    out = args[1] if len(args) > 1 else re.sub(r"\.[^.]+$", "", os.path.basename(src)) + f"_{want}.csv"

    ext = os.path.splitext(src)[1].lower()
    reader = {".csv": rows_from_csv, ".txt": rows_from_csv,
              ".xlsb": rows_from_xlsb, ".xlsx": rows_from_xlsx}.get(ext)
    if not reader:
        raise SystemExit(f"Don't know how to read {ext}. Supported: .csv .txt .xlsb .xlsx")

    scanned = kept = 0
    col = None
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for i, row in enumerate(reader(src)):
            if i == 0:
                col = state_col(row, want)
                if col is None:
                    raise SystemExit("Could not find a state column. Header was:\n  " + str(row[:15]))
                print(f"Filtering on column {col}: {row[col]!r}")
                w.writerow(row)
                continue
            scanned += 1
            if col < len(row) and matches(row[col], want):
                w.writerow(row)
                kept += 1
            if scanned % 500_000 == 0:
                print(f"  ...{scanned:,} rows scanned, {kept:,} kept")

    size = os.path.getsize(out) / 1024
    print(f"\nScanned {scanned:,} rows, kept {kept:,} for {want}.")
    print(f"Wrote {out}  ({size:,.0f} KB)")
    if kept == 0:
        print("\nNothing matched. Check the state column guess above, or pass --state=<abbrev>.")


if __name__ == "__main__":
    main()
